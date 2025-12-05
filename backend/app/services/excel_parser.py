"""
Excel 文件解析器

使用 openpyxl 解析 Excel 文件，提取饮食训练计划数据
"""
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, date
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import re


class ExcelParseError(Exception):
    """Excel 解析错误"""
    pass


class ExcelParser:
    """
    Excel 解析器
    
    功能：
    - 解析 Excel 文件（.xlsx, .xls）
    - 提取日期、餐食、运动、热量、时长等字段
    - 处理不同的 Excel 格式和布局
    - 智能识别列标题
    """
    
    # 列标题关键词映射
    COLUMN_KEYWORDS = {
        'date': ['日期', 'date', '时间', 'time', '日', 'day'],
        'type': ['类型', 'type', '种类', 'category'],
        'name': ['名称', 'name', '项目', 'item'],
        'meal_time': ['餐次', '用餐时间', 'meal', '餐食时间', '早中晚', 'meal time'],
        'food': ['食物', '食品', 'food', '菜品', '餐食内容', '内容'],
        'calories': ['热量', '卡路里', 'calories', 'kcal', '能量', 'calorie'],
        'exercise': ['运动', '锻炼', 'exercise', '训练', '活动'],
        'duration': ['时长', '时间', 'duration', 'time', '分钟', 'min', 'minute'],
        'notes': ['备注', '说明', 'notes', 'remark', '描述', 'note']
    }
    
    def __init__(self):
        """初始化解析器"""
        self.workbook = None
        self.worksheet = None
        self.column_mapping = {}
    
    def parse_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        解析 Excel 文件
        
        Args:
            file_path: Excel 文件路径
            
        Returns:
            解析后的数据列表
            
        Raises:
            ExcelParseError: 解析失败
        """
        try:
            # 打开工作簿
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # 获取活动工作表
            self.worksheet = self.workbook.active
            
            # 识别列标题
            self._identify_columns()
            
            # 提取数据
            data = self._extract_data()
            
            # 关闭工作簿
            self.workbook.close()
            
            return data
            
        except Exception as e:
            if self.workbook:
                self.workbook.close()
            raise ExcelParseError(f"解析 Excel 文件失败: {str(e)}")
    
    def _identify_columns(self) -> None:
        """
        识别列标题
        
        扫描前几行，识别各列的含义
        """
        self.column_mapping = {}
        
        # 扫描前 5 行寻找标题行
        for row_idx in range(1, min(6, self.worksheet.max_row + 1)):
            row = list(self.worksheet.iter_rows(
                min_row=row_idx,
                max_row=row_idx,
                values_only=True
            ))[0]
            
            # 检查这一行是否像标题行
            if self._is_header_row(row):
                self._map_columns(row, row_idx)
                break
        
        # 如果没有找到标题行，使用默认映射
        if not self.column_mapping:
            self._use_default_mapping()
    
    def _is_header_row(self, row: Tuple) -> bool:
        """
        判断是否是标题行
        
        Args:
            row: 行数据
            
        Returns:
            是否是标题行
        """
        if not row:
            return False
        
        # 统计包含关键词的单元格数量
        keyword_count = 0
        non_empty_count = 0
        
        for cell in row:
            if cell:
                non_empty_count += 1
                cell_str = str(cell).lower()
                
                # 检查是否包含任何关键词
                for keywords in self.COLUMN_KEYWORDS.values():
                    if any(keyword.lower() in cell_str for keyword in keywords):
                        keyword_count += 1
                        break
        
        # 如果超过一半的非空单元格包含关键词，认为是标题行
        return non_empty_count > 0 and keyword_count >= non_empty_count * 0.4
    
    def _map_columns(self, header_row: Tuple, row_idx: int) -> None:
        """
        映射列标题到字段名
        
        Args:
            header_row: 标题行数据
            row_idx: 标题行索引
        """
        self.column_mapping = {'header_row': row_idx}
        
        for col_idx, cell in enumerate(header_row, start=1):
            if not cell:
                continue
            
            cell_str = str(cell).lower()
            
            # 匹配字段类型
            for field, keywords in self.COLUMN_KEYWORDS.items():
                if any(keyword.lower() in cell_str for keyword in keywords):
                    self.column_mapping[field] = col_idx
                    break
    
    def _use_default_mapping(self) -> None:
        """使用默认列映射（假设标准格式）"""
        self.column_mapping = {
            'header_row': 1,
            'date': 1,
            'meal_time': 2,
            'food': 3,
            'calories': 4,
            'exercise': 5,
            'duration': 6,
            'notes': 7
        }
    
    def _extract_data(self) -> List[Dict[str, Any]]:
        """
        提取数据
        
        Returns:
            数据列表
        """
        data = []
        header_row = self.column_mapping.get('header_row', 1)
        
        # 从标题行的下一行开始读取数据
        for row_idx in range(header_row + 1, self.worksheet.max_row + 1):
            row_data = self._extract_row(row_idx)
            
            # 跳过空行
            if row_data and any(row_data.values()):
                data.append(row_data)
        
        return data
    
    def _extract_row(self, row_idx: int) -> Dict[str, Any]:
        """
        提取一行数据
        
        Args:
            row_idx: 行索引
            
        Returns:
            行数据字典
        """
        row_data = {}
        
        # 提取各字段
        for field, col_idx in self.column_mapping.items():
            if field == 'header_row':
                continue
            
            cell_value = self.worksheet.cell(row_idx, col_idx).value
            
            # 处理不同字段类型
            if field == 'date':
                row_data[field] = self._parse_date(cell_value)
            elif field == 'calories':
                row_data[field] = self._parse_number(cell_value)
            elif field == 'duration':
                row_data[field] = self._parse_number(cell_value)
            else:
                row_data[field] = self._parse_text(cell_value)
        
        return row_data
    
    def _parse_date(self, value: Any) -> Optional[str]:
        """
        解析日期
        
        Args:
            value: 单元格值
            
        Returns:
            日期字符串（YYYY-MM-DD）或 None
        """
        if not value:
            return None
        
        # 如果已经是 datetime 对象
        if isinstance(value, (datetime, date)):
            return value.strftime('%Y-%m-%d')
        
        # 尝试解析字符串
        if isinstance(value, str):
            # 常见日期格式
            date_patterns = [
                r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})',  # 2024-12-02
                r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',  # 12/02/2024
                r'(\d{4})年(\d{1,2})月(\d{1,2})日',     # 2024年12月02日
            ]
            
            for pattern in date_patterns:
                match = re.search(pattern, value)
                if match:
                    groups = match.groups()
                    
                    # 判断日期格式
                    if len(groups[0]) == 4:  # YYYY-MM-DD
                        year, month, day = groups
                    else:  # MM-DD-YYYY
                        month, day, year = groups
                    
                    try:
                        parsed_date = date(int(year), int(month), int(day))
                        return parsed_date.strftime('%Y-%m-%d')
                    except ValueError:
                        continue
        
        return None
    
    def _parse_number(self, value: Any) -> Optional[float]:
        """
        解析数字
        
        Args:
            value: 单元格值
            
        Returns:
            数字或 None
        """
        if not value:
            return None
        
        # 如果已经是数字
        if isinstance(value, (int, float)):
            return float(value)
        
        # 尝试从字符串中提取数字
        if isinstance(value, str):
            # 移除非数字字符（保留小数点和负号）
            cleaned = re.sub(r'[^\d.-]', '', value)
            
            try:
                return float(cleaned)
            except ValueError:
                return None
        
        return None
    
    def _parse_text(self, value: Any) -> Optional[str]:
        """
        解析文本
        
        Args:
            value: 单元格值
            
        Returns:
            文本或 None
        """
        if not value:
            return None
        
        return str(value).strip()
    
    @staticmethod
    def parse(file_path: str) -> List[Dict[str, Any]]:
        """
        解析 Excel 文件（便捷函数）
        
        Args:
            file_path: Excel 文件路径
            
        Returns:
            解析后的数据列表
        """
        parser = ExcelParser()
        return parser.parse_file(file_path)


# ============================================================================
# 数据转换函数
# ============================================================================

def convert_to_plan_items(parsed_data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """
    将解析后的数据转换为计划项目格式
    
    Args:
        parsed_data: 解析后的原始数据
        
    Returns:
        包含 meals 和 exercises 的字典
    """
    meals = []
    exercises = []
    
    for row in parsed_data:
        # 检查是否有"类型"字段（新格式）
        item_type = row.get('type', '').strip()
        item_name = row.get('name', '').strip()
        
        if item_type and item_name:
            # 新格式：使用"类型"+"名称"
            if item_type in ['餐食', 'meal', '食物']:
                meal = {
                    'date': row.get('date'),
                    'meal_time': row.get('meal_time', '未指定'),
                    'food': item_name,
                    'calories': row.get('calories', 0),
                    'notes': row.get('notes')
                }
                meals.append(meal)
            elif item_type in ['运动', 'exercise', '锻炼']:
                exercise = {
                    'date': row.get('date'),
                    'name': item_name,
                    'duration': row.get('duration', 0),
                    'calories_burned': row.get('calories', 0),
                    'notes': row.get('notes')
                }
                exercises.append(exercise)
        else:
            # 旧格式：使用单独的"食物"和"运动"列
            # 如果有食物信息，添加到餐食列表
            if row.get('food'):
                meal = {
                    'date': row.get('date'),
                    'meal_time': row.get('meal_time', '未指定'),
                    'food': row.get('food'),
                    'calories': row.get('calories', 0),
                    'notes': row.get('notes')
                }
                meals.append(meal)
            
            # 如果有运动信息，添加到运动列表
            if row.get('exercise'):
                exercise = {
                    'date': row.get('date'),
                    'name': row.get('exercise'),
                    'duration': row.get('duration', 0),
                    'calories_burned': row.get('calories', 0),
                    'notes': row.get('notes')
                }
                exercises.append(exercise)
    
    return {
        'meals': meals,
        'exercises': exercises
    }


def group_by_date(plan_items: Dict[str, List[Dict[str, Any]]]) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """
    按日期分组计划项目
    
    Args:
        plan_items: 计划项目（meals 和 exercises）
        
    Returns:
        按日期分组的数据
    """
    grouped = {}
    
    # 分组餐食
    for meal in plan_items.get('meals', []):
        date_key = meal.get('date', 'unknown')
        if date_key not in grouped:
            grouped[date_key] = {'meals': [], 'exercises': []}
        grouped[date_key]['meals'].append(meal)
    
    # 分组运动
    for exercise in plan_items.get('exercises', []):
        date_key = exercise.get('date', 'unknown')
        if date_key not in grouped:
            grouped[date_key] = {'meals': [], 'exercises': []}
        grouped[date_key]['exercises'].append(exercise)
    
    return grouped
